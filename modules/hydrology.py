'''
Created on 04/03/2025

@author: Jose Pedro Matos
'''

import shutil
import pandas as pd
from pathlib import Path
from matplotlib import pyplot as plt
from modules.multi_window_mapper import Multi_Window_Mapper
from modules.quantile_mapping import QuantileMapper, QuantileMapping, QuantileDeltaMapping

# Pour supprimer des avertissements
import warnings
warnings.filterwarnings('ignore')


historical_data_file = Path(r'../data/historical_data.csv')
projections_file = Path(r'../data/projections.csv')

projections = pd.read_csv(projections_file, index_col=0, header = [0, 1, 2, 3], skiprows=[4], parse_dates=[0], date_format='%Y-%m-%d')
historical = pd.read_csv(historical_data_file, index_col=0, header = [0, 1], parse_dates=[0], date_format='%Y-%m-%d')

# Chagement des colonnes pour qu'elles soient similaires à celles de projections
columns = historical.columns.to_frame(index=False)
columns.loc[:, 'Zone'] = columns.loc[:, 'Zone'].map(lambda x: x[0] + x[1:].lower())
columns.loc[:, 'Variables'] = columns.loc[:, 'Variables'].map(lambda x: {'P [mm]': 'pr', 'T [C]': 'tas', 'Hobs [mm]':'hobs', 'Hsim [mm]':'hsim'}[x])
historical.columns = pd.MultiIndex.from_frame(columns)
historical


cas = 'rcp85'
modele = 'CCCma-CanESM2_r1i1p1_SMHI-RCA4_v1'
bv = 'Q18'

def apply_quantile_mapping(cas, modele, bv):
    '''

    '''
    
    tas_historical_bv = historical.loc[:, (bv, 'tas')]
    tas_projection_historical_bv = projections.loc[:, ('tas', 'historical', bv, modele)]
    tas_projection_bv = projections.loc[:, ('tas', cas, bv, modele)]
    tas_projection_bv = pd.concat((tas_projection_bv, tas_projection_historical_bv), axis=1).bfill(axis=1).iloc[:, [0]].sort_index()
    
    pr_historical_bv = historical.loc[:, (bv, 'pr')]
    pr_projection_bv = projections.loc[:, ('pr', cas, bv, modele)]
    pr_projection_historical_bv = projections.loc[:, ('pr', 'historical', bv, modele)]
    pr_projection_bv = pd.concat((pr_projection_bv, pr_projection_historical_bv), axis=1).bfill(axis=1).iloc[:, [0]].sort_index()
    
    # Température
    kw_kernel = {'model': QuantileDeltaMapping, # Le type de Quantile Mapping à utiliser
                 'kw_model': {'trend_window': 10, # la fenêtre pour le calcul du "delta"
                              'transformation': 'additive',
                              'modified': False,
                             },
                }     
    qm = QuantileMapper(projection_historical=tas_projection_bv, reference=tas_historical_bv,
                        kernel=Multi_Window_Mapper, kw_kernel=kw_kernel,
                        trend_window=5, # Le nombre d'années condidérées pour la moyenne glissante
                        hydrological_year_month_start=9)
    qm.map()
    tas_corrected = qm.apply(tas_projection_bv)

    # Précipitation
    kw_kernel = {'model': QuantileMapping, # Le type de Quantile Mapping à utiliser
            } 
    qm = QuantileMapper(projection_historical=pr_projection_bv, reference=pr_historical_bv,
                        kernel=Multi_Window_Mapper, kw_kernel=kw_kernel,
                        trend_window=5, # Le nombre d'années condidérées pour la moyenne glissante
                        hydrological_year_month_start=9)
    qm.map()
    pr_corrected = qm.apply(pr_projection_bv)
    
    return pd.concat((pr_corrected, tas_corrected), keys=['pr', 'tas'], axis=1).loc['1960-09-01':,:]

corrected = apply_quantile_mapping(cas, modele, bv)





hydrological_folder = Path(r'../hydrology/results') / bv / cas / modele
hydrological_folder.mkdir(parents=True, exist_ok=True) # Créer le nouveau dossier

template_hydrological_model = Path(r'../hydrology/template_model.xlsx')
active_hydrological_model = hydrological_folder / 'model.xlsx'

# Faire une copie du modèle de base
_ = shutil.copy(template_hydrological_model, active_hydrological_model)  # For Python 3.8+.

# Trouver le modèle du bassin versant
calibrated_hydrological_model = [i for i in Path(r'../hydrology/base models/').glob(f'MODEL - pluie - débit {bv} *.xlsx')]
if len(calibrated_hydrological_model)!=1:
    raise(Exception(f'Problems with the basin "{bv}". Does the model exist?'))
calibrated_hydrological_model = calibrated_hydrological_model[0]




#===============================================================================
# import xlwings as xw
# 
# try:
#     with xw.Book(calibrated_hydrological_model) as calibrated_model:
#         with xw.Book(active_hydrological_model) as projection_model:
#             calibrated_sheet = calibrated_model.sheets('MODEL - pluie - débit')
#             projection_sheet = projection_model.sheets('MODEL - pluie - débit')
# 
#             projection_sheet.range('G2:N3').value = calibrated_sheet.range('G2:N3').value               # Copier des parametres
#             
#             projection_sheet.range('A6').value = corrected.index[0]                                     # Copier la date de debut
#             projection_sheet[:corrected.shape[0], 6].offset(6,0).value = corrected.iloc[:, 0].values    # copier la precipitation
#             projection_sheet[:corrected.shape[0], 15].offset(6,0).value = corrected.iloc[:, 1].values   # copiear la temperature
#             
#             
# except Exception as ex:
#     raise(ex)
#===============================================================================


import openpyxl
from openpyxl.utils import column_index_from_string

# Load the workbooks and sheets
calibrated_model = openpyxl.load_workbook(calibrated_hydrological_model)
calibrated_sheet = calibrated_model['MODEL - pluie - débit']

projection_model = openpyxl.load_workbook(active_hydrological_model)
projection_sheet = projection_model['MODEL - pluie - débit']

# --- Copier des paramètres ---
# Copy range G2:N3 from calibrated_sheet to projection_sheet.
# In openpyxl, columns and rows are 1-indexed (G is 7, N is 14).
for r_cal, r_proj in zip(
    calibrated_sheet.iter_rows(min_row=2, max_row=3, min_col=7, max_col=14),
    projection_sheet.iter_rows(min_row=2, max_row=3, min_col=7, max_col=14)
):
    for cell_cal, cell_proj in zip(r_cal, r_proj):
        cell_proj.value = cell_cal.value

# --- Copier la date de debut ---
# Write corrected.index[0] to cell A6.
projection_sheet['A6'] = corrected.index[0]

# --- Copier la precipitation ---
# In your xlwings code, you write the precipitation values to a range starting with an offset of 6 rows.
# In openpyxl, we assume that means starting at row 7 (since A6 is row 6).
# Column index 6 in xlwings (0-based) corresponds to column 7 (G) in openpyxl.
start_row_precip = 6  # starting row for precipitation data (G7)
col_precip = 6        # column G
for i, value in enumerate(corrected.iloc[:, 0].values):
    projection_sheet.cell(row=start_row_precip + i, column=col_precip).value = value

# --- Copier la temperature ---
# Similarly, for temperature, xlwings uses column index 15 (0-based) which is column 16 (P) in openpyxl.
start_row_temp = 6  # starting row for temperature data (P7)
col_temp = 17       # column P
for i, value in enumerate(corrected.iloc[:, 1].values):
    projection_sheet.cell(row=start_row_temp + i, column=col_temp).value = value

# Save the updated workbook
projection_model.save(active_hydrological_model)


print('Done!')