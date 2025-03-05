'''
Created on 05/03/2025

@author: Jose Pedro Matos
'''

import shutil
import pandas as pd
import numpy as np
import win32com.client as win32
from pathlib import Path
from matplotlib import pyplot as plt
from modules.multi_window_mapper import Multi_Window_Mapper
from modules.quantile_mapping import QuantileMapper, QuantileMapping, QuantileDeltaMapping

import openpyxl
from openpyxl.utils import column_index_from_string

#===============================================================================
# import warnings
# warnings.filterwarnings('ignore')
#===============================================================================

def apply_quantile_mapping(cas, modele, bv, historical, projections):
    '''

    '''
    
    tas_historical_bv = historical.loc[:, (bv, 'tas')]
    tas_projection_historical_bv = projections.loc[:, ('tas', 'historical', bv, modele)]
    tas_projection_bv = projections.loc[:, ('tas', cas, bv, modele)]
    tas_projection_bv = pd.concat((tas_projection_bv, tas_projection_historical_bv), axis=1).bfill(axis=1).iloc[:, [0]].sort_index()
    
    pr_historical_bv = historical.loc[:, (bv, 'pr')]
    pr_projection_bv = projections.loc[:, ('pr', cas, bv, modele)]
    pr_projection_historical_bv = projections.loc[:, ('pr', 'historical', bv, modele)]
    pr_projection_bv = pd.concat((pr_projection_bv, pr_projection_historical_bv), axis=1).bfill(axis=1).iloc[:, [0]].sort_index()
    
    # Température
    kw_kernel = {'model': QuantileDeltaMapping, # Le type de Quantile Mapping à utiliser
                 'kw_model': {'trend_window': 10, # la fenêtre pour le calcul du "delta"
                              'transformation': 'additive',
                              'modified': False,
                             },
                }     
    qm = QuantileMapper(projection_historical=tas_projection_bv, reference=tas_historical_bv,
                        kernel=Multi_Window_Mapper, kw_kernel=kw_kernel,
                        trend_window=5, # Le nombre d'années condidérées pour la moyenne glissante
                        hydrological_year_month_start=9)
    qm.map()
    tas_corrected = qm.apply(tas_projection_bv)

    # Précipitation
    kw_kernel = {'model': QuantileMapping, # Le type de Quantile Mapping à utiliser
            } 
    qm = QuantileMapper(projection_historical=pr_projection_bv, reference=pr_historical_bv,
                        kernel=Multi_Window_Mapper, kw_kernel=kw_kernel,
                        trend_window=5, # Le nombre d'années condidérées pour la moyenne glissante
                        hydrological_year_month_start=9)
    qm.map()
    pr_corrected = qm.apply(pr_projection_bv)
    
    return pd.concat((pr_corrected, tas_corrected), keys=['pr', 'tas'], axis=1).loc['1960-09-01':,:]


def excel_write(cas, modele, bv, corrected,
                       calibrated_hydrological_model,
                       results_path=Path(r'hydrology/results'),
                       template_hydrological_model=Path(r'hydrology/template_model.xlsx'),
                       sheet_name='MODEL - pluie - débit',
                      ):
    '''
    
    '''

    hydrological_folder = results_path / bv / cas / modele
    hydrological_folder.mkdir(parents=True, exist_ok=True) # Créer le nouveau dossier
    active_hydrological_model = hydrological_folder / 'model.xlsx'

    # Faire une copie du modèle de base
    _ = shutil.copy(template_hydrological_model, active_hydrological_model)  # For Python 3.8+.
    
    # Load the workbooks and sheets
    calibrated_model = openpyxl.load_workbook(calibrated_hydrological_model)
    calibrated_sheet = calibrated_model[sheet_name]

    projection_model = openpyxl.load_workbook(active_hydrological_model)
    projection_sheet = projection_model[sheet_name]

    # --- Copier des paramètres ---
    # Copy range G2:N3 from calibrated_sheet to projection_sheet.
    # In openpyxl, columns and rows are 1-indexed (G is 7, N is 14).
    for r_cal, r_proj in zip(
        calibrated_sheet.iter_rows(min_row=2, max_row=3, min_col=7, max_col=14),
        projection_sheet.iter_rows(min_row=2, max_row=3, min_col=7, max_col=14)
    ):
        for cell_cal, cell_proj in zip(r_cal, r_proj):
            cell_proj.value = cell_cal.value

    # --- Copier la date de debut ---
    # Write corrected.index[0] to cell A6.
    projection_sheet['A6'] = corrected.index[0]

    # --- Copier la precipitation ---
    # In your xlwings code, you write the precipitation values to a range.
    start_row_precip = 6  
    col_precip = 6       
    for i, value in enumerate(corrected.iloc[:, 0].values):
        projection_sheet.cell(row=start_row_precip + i, column=col_precip).value = value

    # --- Copier la temperature ---
    start_row_temp = 6 
    col_temp = 17      
    for i, value in enumerate(corrected.iloc[:, 1].values):
        projection_sheet.cell(row=start_row_temp + i, column=col_temp).value = value

    # Save the updated workbook
    projection_model.save(active_hydrological_model)
                          
    print(f'Model created at: {active_hydrological_model}')
    
    return active_hydrological_model

def excel_read(active_hydrological_model, sheet_name='MODEL - pluie - débit', verbose=1):
    '''
     
    '''
 
    projection_model = openpyxl.load_workbook(active_hydrological_model, data_only=True)
    projection_sheet = projection_model[sheet_name]
 
    simulation = pd.DataFrame(np.empty((12*200, 1))) * pd.NA
    simulation.columns = ['hsim']
     
    start_row_temp = 6  
    col_temp = 15       
    for i0 in range(simulation.shape[0]):
        simulation.iloc[i0, 0] = projection_sheet.cell(row=start_row_temp + i0, column=col_temp).value
    simulation = simulation.dropna()
         
    start_date = projection_sheet.cell(row=6, column=1).value
    simulation.index = pd.date_range(start=start_date, periods=simulation.shape[0], freq='MS')
         
    projection_model.close()
             
    if verbose>=1:
        print(f'Model created at: {active_hydrological_model}')
     
    return simulation


def run_excel(workbook_path):
    
    workbook_path = str(workbook_path.absolute())
    
    # Open Excel in the background
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel.Visible = False

    # Open the workbook
    wb = excel.Workbooks.Open(workbook_path)

    # Refresh any external data if needed
    wb.RefreshAll()

    # Force Excel to recalculate everything using the Application object
    excel.CalculateFull()

    # Save and close the workbook
    wb.Save()
    wb.Close()
    excel.Quit()
    
    print('Excel calculated')
    


def bb():
    pass

pass